from django.shortcuts import render, redirect
from .models import LabDetails, StudentsDetails,Availability,ExamAllocation
from .forms import AvailabilityForm,ExcelUploadForm
from django.utils import timezone
from datetime import date
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.core.paginator import Paginator
import json
import string
import random
from io import BytesIO
import xlwt
import openpyxl
from django.http import HttpResponse
from django.shortcuts import render, get_object_or_404
from .models import ExamAllocation,CredentialsManager
from .forms import CredentialsManagerForm
from openpyxl import Workbook
from django.core.mail import EmailMessage
from .models import GeneratedCredential
from django.contrib import messages
from django.urls import reverse
from django.utils.timezone import now
from django.http import FileResponse
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4, landscape
import re
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
from django.utils.dateformat import format as django_date_format
from django.contrib.auth.decorators import login_required
from faculty.models import LabRequest
from .models import AssignedCourses
from GCR_draft.models import CourseEntry
import pandas as pd

@login_required
def faculty_dashboard(request):
    lab_records = LabDetails.objects.all()
    dashboard_data = []

    for lab in lab_records:
        student_count = lab.students.count()  

        dashboard_data.append({
            'course_code': lab.course_code,
            'course_name': lab.course_name,
            'batch_no': lab.batch_no,
            'faculty_name': lab.faculty_name,  
            'student_strength': student_count,
            'lab_no': lab.ctc_lab_no,
            'day': lab.day,
            'hours': lab.hours,
        })

    return render(request, 'faculty_dashboard.html', {'dashboard_data': dashboard_data})





@login_required
def allocate_exam_view(request, course_code, course_name, batch_no):
    if request.method == 'POST':
        selected_date = request.POST.get('date')
        session = request.POST.get('session')
        lab_venue = request.POST.get('lab_venue')
        print(f"Allocated: {course_code} - {course_name} - Batch {batch_no}")
        print(f"Date: {selected_date}, Session: {session}, Lab: {lab_venue}")
        return redirect('faculty_dashboard')

    today = date.today().isoformat()

    labs = []
    for i in range(1, 19):
        block = (i - 1) // 6 + 1
        labs.append((i, f"CTC-{block}:Lab {i}"))

    context = {
        'course_code': course_code,
        'course_name': course_name,
        'batch_no': batch_no,
        'min_date': today,
        'labs': labs,
    }
    return render(request, 'allocating_form.html', context)

@login_required
def check_availability(request):
    if request.method == "POST":
        print("POST DATA:", request.POST)
        date = request.POST.get("date")
        session = request.POST.get("session")
        venue = request.POST.get("lab_venue")
        course_code = request.POST.get("course_code")
        course_name = request.POST.get("course_name")
        batch_no = request.POST.get("batch_no")

        try:
            lab = LabDetails.objects.get(course_code=course_code, course_name=course_name, batch_no=batch_no)
        except LabDetails.DoesNotExist:
            return JsonResponse({"status": "Lab Not Found"})

        exists = Availability.objects.filter(date=date, session=session, lab_venue=venue).exists()

        if exists:
            return JsonResponse({"status": "Not Available"})
        else:
            Availability.objects.create(
                lab_details=lab,
                date=date,
                session=session,
                lab_venue=venue,
                status=True
            )
            return JsonResponse({"status": "Available"})


@login_required
@csrf_exempt
def confirm_allocation(request):
    if request.method == "POST":
        print("DEBUG POST:", request.POST)

        date = request.POST.get("date")
        session = request.POST.get("session")
        venue = request.POST.get("lab_venue")
        course_code = request.POST.get("course_code")
        course_name = request.POST.get("course_name")
        batch_no = request.POST.get("batch_no")

        try:
            lab_details = LabDetails.objects.get(
                course_code=course_code,
                course_name=course_name,
                batch_no=batch_no
            )
        except LabDetails.DoesNotExist:
            return JsonResponse({"success": False, "error": "Lab details not found"})
        student_strength = lab_details.students.count()

    
        Availability.objects.get_or_create(
            lab_details=lab_details,
            date=date,
            session=session,
            lab_venue=venue,
            defaults={"status": True}
        )


        ExamAllocation.objects.create(
            lab_details=lab_details,
            date=date,
            session=session,
            lab_venue=venue,
            student_strength=student_strength
        )

        return JsonResponse({"success": True})
    

@login_required
def course_allotment_view(request):
    allocations_list = ExamAllocation.objects.select_related('lab_details').order_by('-date')
    paginator = Paginator(allocations_list, 10)  

    page_number = request.GET.get('page')
    allocations = paginator.get_page(page_number)

    return render(request, 'course_allotment.html', {'allocations': allocations})

@login_required
def credentials_generation_view(request):
    allocations = ExamAllocation.objects.select_related('lab_details').order_by('-date')
    return render(request, 'credentials_generation.html', {'allocations': allocations})

@login_required
def generate_credentials_excel(request, allocation_id):
    allocation = get_object_or_404(ExamAllocation, id=allocation_id)
    students = allocation.lab_details.students.all()

    course_name = allocation.lab_details.course_name.upper().replace(" ", "")
    prefix = course_name[:2] + "L" + str(allocation.lab_details.batch_no)
    prefix = prefix[:5]  

    # Exclude confusing characters
    characters = ''.join(set(string.ascii_uppercase + string.digits) - set("oOiILl10"))

    response = HttpResponse(content_type='application/ms-excel')
    filename = f"Credentials_{course_name}.xls"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    wb = xlwt.Workbook(encoding='utf-8')
    ws = wb.add_sheet('Credentials')

    ws.write(0, 0, 'S.No')
    ws.write(0, 1, 'Username')
    ws.write(0, 2, 'Password')

    for i, student in enumerate(students, start=1):
        username = f"{prefix}{i:02}"
        password = ''.join(random.choices(characters, k=6))

        ws.write(i, 0, i)
        ws.write(i, 1, username)
        ws.write(i, 2, password)

    wb.save(response)
    return response


@login_required
def credentials_manager_view(request): 
    credentials = CredentialsManager.objects.all()

    faculty_ids = credentials.values_list('ctc_faculty_id', flat=True).distinct()
    batch_nos = credentials.values_list('batch_no', flat=True).distinct()
    dates = credentials.values_list('date', flat=True).distinct()
    sessions = credentials.values_list('session', flat=True).distinct()

    context = {
        'credentials': credentials,
        'faculty_ids': faculty_ids,
        'batch_nos': batch_nos,
        'dates': dates,
        'sessions': sessions,
    }
    return render(request, 'credentials_manager.html', context)

@login_required
@csrf_exempt
def generate_credentials(request, cred_id):
    if request.method == 'POST':
        try:
            cm = CredentialsManager.objects.get(id=cred_id)
        except CredentialsManager.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Invalid credential ID'})

        if GeneratedCredential.objects.filter(
            course_code=cm.course_code,
            batch_no=cm.batch_no,
            date=cm.date,
            session=cm.session,
            lab_venue=cm.lab_venue
        ).exists():
            return JsonResponse({'status': 'warning', 'message': 'Credentials already exist!'})

        batch_number_match = re.search(r'\d+', cm.batch_no)
        batch_number = batch_number_match.group() if batch_number_match else "1"

        password_chars = "ABCDEFGHJKMNPQRSTUVWXYZ"

        for i in range(1, 51):
            serial = str(i).zfill(2)
            username = f"{cm.session_no}{serial}"
            password = ''.join(random.choices(password_chars, k=6))

            GeneratedCredential.objects.create(
                course_code=cm.course_code,
                course_name=cm.course_name,
                batch_no=cm.batch_no,
                date=cm.date,
                session=cm.session,
                lab_venue=cm.lab_venue,
                username=username,
                password=password,
                regno=''
            )

        cm.last_generated = now()
        cm.save(update_fields=['last_generated'])

        return JsonResponse({
            'status': 'success',
            'message': 'Credentials generated successfully!',
            'last_generated': cm.last_generated.strftime('%d-%m-%Y %H:%M:%S')  
        })

    return JsonResponse({'status': 'error', 'message': 'Invalid request method'})



@login_required
@csrf_exempt
def send_credentials_email(request, cred_id):
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Invalid request method'})

    try:
        cm = CredentialsManager.objects.get(id=cred_id)
    except CredentialsManager.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Invalid credential ID'})

    credentials = GeneratedCredential.objects.filter(
        course_code=cm.course_code,
        batch_no=cm.batch_no,
        date=cm.date,
        session=cm.session,
        lab_venue=cm.lab_venue,
    )

    if not credentials.exists():
        return JsonResponse({'status': 'error', 'message': 'No credentials found for this record'})

    # Check recipient email
    if not cm.ctc_faculty_mail_id:
        return JsonResponse({'status': 'error', 'message': 'Faculty email not found'})

    try:
        # PDF Generation
        pdf_file = BytesIO()
        doc = SimpleDocTemplate(pdf_file, pagesize=A4)
        elements = []

        styles = getSampleStyleSheet()
        elements.append(Paragraph("Lab Exam Credentials", styles['Heading1']))
        elements.append(Spacer(1, 12))

        formatted_date = cm.date.strftime('%d-%m-%Y')
        meta_info = f"Venue: {cm.lab_venue}<br/>Session: {cm.session}<br/>Date: {formatted_date}<br/>Course: {cm.course_code} - {cm.course_name} ({cm.batch_no})<br/>Total Credentials: {credentials.count()}"
        elements.append(Paragraph(meta_info, styles['Normal']))
        elements.append(Spacer(1, 20))

        data = [['Username', 'Password', 'Regno']] + [
            [cred.username, cred.password, cred.regno or ''] for cred in credentials
        ]

        table = Table(data, colWidths=[150, 150, 150])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ]))

        elements.append(table)
        elements.append(Spacer(1, 30))
        elements.append(Paragraph("Name of the Faculty:<br/><br/>ID:<br/><br/>Signature of the Faculty:", styles['Normal']))

        doc.build(elements)
        pdf_file.seek(0)

        # Email setup
        email = EmailMessage(
            subject='Lab Exam Credentials (PDF)',
            body=(
                f"Dear Staff,\n\n"
                f"Please find the attached PDF file containing the credentials for the upcoming lab exam.\n\n"
                f"Venue: {cm.lab_venue} | {formatted_date} | {cm.session}\n"
                f"Course: {cm.course_code} - {cm.course_name} ({cm.batch_no})\n\n"
                f"Regards,\nCTC, KITS."
            ),
            from_email='joneabishek@karunya.edu',
            to=[cm.ctc_faculty_mail_id],
        )

        email.attach(
            f'Credentials_{cm.course_code}_{cm.batch_no}.pdf',
            pdf_file.getvalue(),
            'application/pdf'
        )

        email.send()

        return JsonResponse({'status': 'success', 'message': 'PDF emailed successfully!'})

    except Exception as e:
        print("EMAIL SEND ERROR:", str(e))
        return JsonResponse({'status': 'error', 'message': f'Email sending failed: {str(e)}'})
    
@login_required   
def download_credentials_excel(request, cred_id):
    try:
        cm = CredentialsManager.objects.get(id=cred_id)
    except CredentialsManager.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Invalid ID'})

    credentials = GeneratedCredential.objects.filter(
        course_code=cm.course_code,
        batch_no=cm.batch_no,
        date=cm.date,
        session=cm.session,
        lab_venue=cm.lab_venue,
    )

    formatted_date = cm.date.strftime('%d-%m-%Y')  

    wb = Workbook()
    ws = wb.active

    ws.append([f"{cm.lab_venue} : {cm.session} - {formatted_date}"])  
    ws.append([f"{cm.course_code} : {cm.course_name} - {cm.batch_no}"])
    ws.append(["Total Credentials: 80"])
    ws.append(['Username', 'Password', 'Regno'])

    for cred in credentials:
        ws.append([cred.username, cred.password, cred.regno or ''])

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"Credentials_{cm.course_code}_{cm.batch_no}.xlsx"
    return FileResponse(buffer, as_attachment=True, filename=filename)






@login_required
def download_all_credentials_pdf(request):
    pdf_buffer = BytesIO()
    doc = SimpleDocTemplate(
    pdf_buffer,
    pagesize=landscape(A4),
    leftMargin=1 * inch,     
    rightMargin=0.7 * inch,
    topMargin=0.5 * inch,     
    bottomMargin=0.5 * inch
    )

    elements = []
    styles = getSampleStyleSheet()

    elements.append(Paragraph("All Generated Credentials", styles['Heading1']))
    elements.append(Spacer(1, 12))

    headers = ["S.No", "Course Code", "Course Name", "Batch", "Date", "Session", "Venue", "Username", "Password", "Regno"]
    data = [headers]

    creds = GeneratedCredential.objects.all()
    for i, cred in enumerate(creds, start=1):
        row = [
            str(i),
            cred.course_code,
            cred.course_name,
            cred.batch_no,
            cred.date.strftime('%d-%m-%Y'),
            cred.session,
            cred.lab_venue,
            cred.username,
            cred.password,
            cred.regno or "-"
        ]
        data.append(row)

    table = Table(data, repeatRows=1, colWidths=[
        0.4 * inch, 0.9 * inch,  2.9 * inch, 0.9 * inch, 1 * inch,
        1.5 * inch, 1.3 * inch, 1 * inch, 1 * inch, 1 * inch
    ])

    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#cccccc')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        
    ]))
    elements.append(table)
    doc.build(elements)
    pdf_buffer.seek(0)

    excel_buffer = BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Credentials"

    ws.append(headers)  
    for i, cred in enumerate(creds, start=1):
        ws.append([
            i,
            cred.course_code,
            cred.course_name,
            cred.batch_no,
            cred.date.strftime('%d-%m-%Y'),
            cred.session,
            cred.lab_venue,
            cred.username,
            cred.password,
            cred.regno or "-"
        ])
    wb.save(excel_buffer)
    excel_buffer.seek(0)

    email = EmailMessage(
        subject="All Generated Credentials",
        body="Please find attached the credentials in PDF and Excel formats.",
        from_email="joneabishek@karunya.edu",
        to=["johnabishek86@gmail.com"]
    )
    email.attach('credentials.pdf', pdf_buffer.getvalue(), 'application/pdf')
    email.attach('credentials.xlsx', excel_buffer.getvalue(), 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    email.send()

    # pdf_buffer.seek(0)
    return FileResponse(pdf_buffer, as_attachment=True, filename='All_Credentials.pdf')




@login_required
def upload_excel(request):
    if request.method == 'POST':
        form = ExcelUploadForm(request.POST, request.FILES)
        if form.is_valid():
            file = request.FILES['file']
            wb = openpyxl.load_workbook(file)
            sheet = wb.active

            
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not any(row):  # Skip empty rows
                    continue

                try:
                    if len(row) < 10:
                        messages.warning(request, f"Skipped row due to not enough columns: {row}")
                        continue
                    row = row[:10]

                    session_no, date, session, course_code, course_name, batch_no, lab_venue, faculty_name, faculty_id, faculty_email = row

                    # Check for missing required fields
                    missing_fields = []
                    if not course_code: missing_fields.append("course_code")
                    if not batch_no: missing_fields.append("batch_no")
                    if not date: missing_fields.append("date")

                    if missing_fields:
                        print(f"Row skipped due to missing fields {missing_fields}: {row}")
                        messages.warning(request, f"Skipped row due to missing fields: {', '.join(missing_fields)}")
                        continue

                    faculty_id = str(faculty_id).strip()
                    if faculty_id.endswith('.0'):
                        faculty_id = faculty_id[:-2]


                    CredentialsManager.objects.create(
                        session_no=session_no,
                        course_code=course_code,
                        course_name=course_name,
                        batch_no=batch_no,
                        date=date,
                        session=session,
                        lab_venue=lab_venue,
                        ctc_faculty_id=faculty_id,
                        ctc_faculty_name=faculty_name,
                        ctc_faculty_mail_id=faculty_email
                    )

                except Exception as e:
                    print(f"Error processing row: {row} -> {e}")
                    messages.error(request, f"Error processing a row: {e}")

            messages.success(request, 'Excel uploaded successfully!')
            return redirect('upload_excel')

    else:
        form = ExcelUploadForm()
    return render(request, 'upload_excel.html', {'form': form})


@login_required
def division_requests(request):
    lab_requests = LabRequest.objects.all()
    assigned_courses = AssignedCourses.objects.all()

    
    assigned_map = {
        course.course_code: course.assigned_date for course in assigned_courses
    }

    for req in lab_requests:
        req.assigned_date = assigned_map.get(req.course_code)

    return render(request, 'division_requests.html', {
        'lab_requests': lab_requests
    })


@login_required
def assign_course(request, lab_id):
    lab_req = LabRequest.objects.get(id=lab_id)

    AssignedCourses.objects.create(
        course_code=lab_req.course_code,
        course_name=lab_req.course_name,
        department_name=lab_req.department_name,
        semester=lab_req.semester,
        year=lab_req.year,
        degree=lab_req.degree,
        programme=lab_req.programme,
        numberofcredits=lab_req.numberofcredits,
        num_slots=lab_req.num_slots,
        num_batches=lab_req.num_batches,
        os_required=lab_req.os_required,
        software_required_version_opensource=lab_req.software_required_version_opensource,
        software_required_version_license=lab_req.software_required_version_license,
        remarks=lab_req.remarks
    )
    messages.success(request, "Lab has been assigned successfully.")
    return redirect('division_requests')




@login_required
def download_assigned_courses(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Assigned Courses"

    headers = [
        "Department", "Semester", "Year", "Degree", "Programme", "Number of Credits",
        "Course Code", "Course Name", "No of Slots", "No of Batches", "OS Required",
        "Software(Open Source)", "Software(License)", "Remarks", "Assigned Date"
    ]
    ws.append(headers)

    for req in AssignedCourses.objects.all():
        ws.append([
            req.department_name, req.semester, req.year, req.degree, req.programme,
            req.numberofcredits, req.course_code, req.course_name, req.num_slots,
            req.num_batches, req.os_required, req.software_required_version_opensource,
            req.software_required_version_license, req.remarks,
            req.assigned_date.strftime("%Y-%m-%d") if req.assigned_date else ''
        ])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=assigned_courses.xlsx'
    wb.save(response)
    return response









# Google classroom work

def course_entry_list(request):
    courses = CourseEntry.objects.all()
    return render(request, 'course_entry_list.html', {'courses': courses})



def download_course_entries_excel(request):
    queryset = CourseEntry.objects.all().values()
    df = pd.DataFrame(list(queryset))

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=course_entries.xlsx'

    with pd.ExcelWriter(response, engine='xlsxwriter') as writer:
        df.to_excel(writer, index=False, sheet_name='CourseEntries')

    return response

